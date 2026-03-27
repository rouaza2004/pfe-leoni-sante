from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from medical.models import StockItem

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - runtime dependency
    raise CommandError(
        "openpyxl is required. Install it with: pip install openpyxl"
    ) from exc


STOP_WORDS = {
    "medicaments",
    "medicament",
    "medic",
    "autres",
    "matricule",
    "quantite",
    "consultation",
    "classement",
    "soins",
}


def _normalize(text: str) -> str:
    return re.sub(r"\\s+", " ", text.strip().lower())


def _has_medic_keyword(text: str) -> bool:
    return re.search(r"medic|m\u00e9dic", text, re.IGNORECASE) is not None


class Command(BaseCommand):
    help = "Import medications list from an Excel file into StockItem (type MEDICAMENT)."

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            type=str,
            help="Path to the Excel file (e.g. C:\\\\Users\\\\...\\\\DOC (6).xlsx).",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Optional sheet name to scan (default: scan all sheets).",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).expanduser()
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheets = [options["sheet"]] if options["sheet"] else wb.sheetnames

        meds = []

        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"Sheet not found: {sheet_name}"))
                continue

            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    if not _has_medic_keyword(value):
                        continue

                    row_idx = cell.row + 1
                    for next_cell in ws[row_idx]:
                        if not isinstance(next_cell.value, str):
                            continue
                        text = _normalize(next_cell.value)
                        if not text:
                            continue
                        if any(stop in text for stop in STOP_WORDS):
                            continue
                        meds.append(next_cell.value.strip())
                    break
                if meds:
                    break
            if meds:
                break

        if not meds:
            raise CommandError("No medications found in the file.")

        created = 0
        skipped = 0

        with transaction.atomic():
            for name in meds:
                exists = StockItem.objects.filter(nom__iexact=name).exists()
                if exists:
                    skipped += 1
                    continue
                StockItem.objects.create(
                    nom=name,
                    type_article="MEDICAMENT",
                    quantite=0,
                    seuil_critique=0,
                    unite="unite",
                )
                created += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported medications. Created: {created}, Skipped (existing): {skipped}"
            )
        )
